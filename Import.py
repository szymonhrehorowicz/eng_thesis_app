# This Python file uses the following encoding: utf-8
from PySide6.QtCore import Slot, QCoreApplication
from PySide6.QtWidgets import QFileDialog
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from utilities import error_dialog, success_dialog
import os

class Import:
    def __init__(self, handler):
        self.handler = handler
        self.filename = ""

    @Slot()
    def load_file(self):
        self.filename = QFileDialog.getOpenFileName(None, QCoreApplication.tr("Wybierz plik z nastawami regulatorow"), filter="Excel files (*.xlsx)")[0]

        try:
            workbook = load_workbook(self.filename)
        except InvalidFileException:
            error_dialog(self.handler, QCoreApplication.tr("Wybrano zły plik"))
            return
        
        sheet = workbook.active
        idx_start = 3
        student_idxs = []

        for student_entry in tuple(sheet.columns)[0][idx_start:]:
            student_idx = student_entry.value
            if student_idx:
                student_idxs.append(student_idx)
                self.handler.ui.cmbImportIdx.addItem(str(student_idx))
            
        if len(student_idxs) == 0:
            self.handler.ui.btnImportAction.setEnabled(False)
            error_dialog(self.handler, QCoreApplication.tr("W pliku nie ma żadnych poprawnych danych do wczytania"))
        else:
            self.handler.ui.btnImportAction.setEnabled(True)
            success_dialog(self.handler, QCoreApplication.tr("Dane wczytano pomyślnie"))

    @Slot()
    def import_data(self):
        student_idx = self.handler.ui.cmbImportIdx.currentText()

        try:
            workbook = load_workbook(self.filename)
        except InvalidFileException:
            error_dialog(self.handler, QCoreApplication.tr("Wybrano zły plik"))
            return
        
        sheet = workbook.active
        idx_start = 3
        row_of_interest = 0

        for row_idx, student_entry in enumerate(tuple(sheet.columns)[0][idx_start:]):
            if str(student_entry.value) == student_idx:
                row_of_interest = row_idx + idx_start
        
        if row_of_interest == 0:
            error_dialog(self.handler, QCoreApplication.tr("W pliku nie ma danych studenta z podanym numerem albumu"))
            return
        
        params = [elem.value for elem in tuple(sheet.rows)[row_of_interest][1:]]
        self._update_parameters(params)

    _INDICES = {
        "heater_bb_x": 0,
        "heater_bb_hysteresis": 1,
        "heater_bb_power": 2,
        "heater_pid_x": 3,
        "heater_pid_kp": 4,
        "heater_pid_ki": 5,
        "heater_pid_kd": 6,
        "heater_pid_ti": 7,
        "heater_pid_td": 8,
        "heater_pid_kaw": 9,
        "heater_coil_ref": 10,
        "heater_temp_ref": 11,
        "fan_bb_x": 12,
        "fan_bb_hysteresis": 13,
        "fan_pid_x": 14,
        "fan_pid_kp": 15,
        "fan_pid_ki": 16,
        "fan_pid_kd": 17,
        "fan_pid_ti": 18,
        "fan_pid_td": 19,
        "fan_pid_kaw": 20
    }

    def _update_parameters(self, params):
        # Heater BB
        if params[self._INDICES["heater_bb_x"]]:
            self.handler.ui.inHeaterBBSetValue.setValue(int(params[self._INDICES["heater_bb_x"]]))
            self.handler.heaterControlsHandler.bb_setValue()
        if params[self._INDICES["heater_bb_hysteresis"]]:
            self.handler.ui.inHeaterBBHysteresis.setValue(int(params[self._INDICES["heater_bb_hysteresis"]]))
            self.handler.heaterControlsHandler.bb_setHysteresis()
        if params[self._INDICES["heater_bb_power"]]:
            self.handler.ui.inHeaterBBPower.setValue(int(params[self._INDICES["heater_bb_power"]]))
            self.handler.heaterControlsHandler.bb_setPower(int(params[self._INDICES["heater_bb_power"]]))
        # Heater PID
        if params[self._INDICES["heater_pid_x"]]:
            self.handler.ui.inHeaterPIDSetValue.setValue(int(params[self._INDICES["heater_pid_x"]]))
            self.handler.heaterControlsHandler.pid_setValue()
        if params[self._INDICES["heater_pid_kp"]]:
            self.handler.ui.inHeaterPID_Kp.setValue(float(params[self._INDICES["heater_pid_kp"]]))
            self.handler.heaterControlsHandler.pid_setKp()
        if params[self._INDICES["heater_pid_ki"]]:
            self.handler.ui.inHeaterPID_Ki.setValue(float(params[self._INDICES["heater_pid_ki"]]))
            self.handler.heaterControlsHandler.pid_setKi()
        if params[self._INDICES["heater_pid_kd"]]:
            self.handler.ui.inHeaterPID_Kd.setValue(float(params[self._INDICES["heater_pid_kd"]]))
            self.handler.heaterControlsHandler.pid_setKd()
        if params[self._INDICES["heater_pid_ti"]]:
            self.handler.ui.inHeaterPID_Ti.setValue(float(params[self._INDICES["heater_pid_ti"]]))
            self.handler.heaterControlsHandler.pid_setTi()
        if params[self._INDICES["heater_pid_td"]]:
            self.handler.ui.inHeaterPID_Td.setValue(float(params[self._INDICES["heater_pid_td"]]))
            self.handler.heaterControlsHandler.pid_setTd()
        if params[self._INDICES["heater_pid_kaw"]]:
            self.handler.ui.inHeaterPID_Kaw.setValue(float(params[self._INDICES["heater_pid_kaw"]]))
            self.handler.heaterControlsHandler.pid_setKaw()
        if params[self._INDICES["heater_coil_ref"]]:
            print(params[self._INDICES["heater_coil_ref"]])
            if int(params[self._INDICES["heater_coil_ref"]]) == 17:
                self.handler.mainMenuHandler.set_heater_high_power()
            else:
                self.handler.mainMenuHandler.set_heater_low_power()
        if params[self._INDICES["heater_temp_ref"]]:
            if params[self._INDICES["heater_temp_ref"]] == "lewa":
                self.handler.ui.btnHeaterControllerSetRightCoil.setChecked(False)
                self.handler.ui.btnHeaterControllerSetLeftCoil.setChecked(True)
            else:
                self.handler.ui.btnHeaterControllerSetRightCoil.setChecked(True)
                self.handler.ui.btnHeaterControllerSetLeftCoil.setChecked(False)
        # Fan BB
        if params[self._INDICES["fan_bb_x"]]:
            self.handler.ui.inFanBBSetValue.setValue(int(params[self._INDICES["fan_bb_x"]]))
            self.handler.fanControlsHandler.bb_setValue()
        if params[self._INDICES["fan_bb_hysteresis"]]:
            self.handler.ui.inFanBBHysteresis.setValue(int(params[self._INDICES["fan_bb_hysteresis"]]))
            self.handler.fanControlsHandler.bb_setHysteresis()
        # Fan PID
        if params[self._INDICES["fan_pid_x"]]:
            self.handler.ui.inFanPIDSetValue.setValue(int(params[self._INDICES["fan_pid_x"]]))
            self.handler.fanControlsHandler.pid_setValue()
        if params[self._INDICES["fan_pid_kp"]]:
            self.handler.ui.inFanPID_Kp.setValue(float(params[self._INDICES["fan_pid_kp"]]))
            self.handler.fanControlsHandler.pid_setKp()
        if params[self._INDICES["fan_pid_ki"]]:
            self.handler.ui.inFanPID_Ki.setValue(float(params[self._INDICES["fan_pid_ki"]]))
            self.handler.fanControlsHandler.pid_setKi()
        if params[self._INDICES["fan_pid_kd"]]:
            self.handler.ui.inFanPID_Kd.setValue(float(params[self._INDICES["fan_pid_kd"]]))
            self.handler.fanControlsHandler.pid_setKd()
        if params[self._INDICES["fan_pid_ti"]]:
            self.handler.ui.inFanPID_Ti.setValue(float(params[self._INDICES["fan_pid_ti"]]))
            self.handler.fanControlsHandler.pid_setTi()
        if params[self._INDICES["fan_pid_td"]]:
            self.handler.ui.inFanPID_Td.setValue(float(params[self._INDICES["fan_pid_td"]]))
            self.handler.fanControlsHandler.pid_setTd()
        if params[self._INDICES["fan_pid_kaw"]]:
            self.handler.ui.inFanPID_Kaw.setValue(float(params[self._INDICES["fan_pid_kaw"]]))
            self.handler.fanControlsHandler.pid_setKaw()

